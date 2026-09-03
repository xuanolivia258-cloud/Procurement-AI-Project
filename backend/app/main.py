import io
import json
import math
import re
import time
from contextlib import asynccontextmanager
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal, InvalidOperation
from html import escape
from uuid import uuid4

from fastapi import Depends, FastAPI, HTTPException, Query, Request, Response
from fastapi.exceptions import RequestValidationError
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.data_source import AxDataSource, StrData, StrRef, StrVal
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Font, PatternFill
import httpx
from sqlalchemy import case, func, or_, select, update
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from .auth import get_actor
from .config import settings
from .database import Base, SessionLocal, engine, get_db
from .logging_config import (
    log_access, log_integration_event, log_operation, log_request_error, summarize_http_response,
)
from .models import Lifecycle, Project, ProjectAuditLog, ReferenceOption, utcnow
from .schemas import (
    Actor, AuditLogRead, BulkProjectDeleteRequest, Currency, LifecycleRequest, PaginatedProjects, ProjectCreate,
    ProjectRead, ProjectUpdate, ReferenceOptionCreate, ReferenceOptionRead,
    ReferenceOptionUpdate,
)


@asynccontextmanager
async def lifespan(_app: FastAPI):
    Base.metadata.create_all(bind=engine)
    with SessionLocal() as db:
        seed_reference_options(db)
    log_operation("service_start", "system", count=0, message=f"Backend service started; auth_mode={settings.auth_mode}.")
    yield


app = FastAPI(title=settings.app_name, version="2.0.0", lifespan=lifespan)
app.add_middleware(
    CORSMiddleware,
    allow_origins=settings.cors_origin_list,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.middleware("http")
async def request_id_middleware(request: Request, call_next):
    started = time.perf_counter()
    request_id = request.headers.get("X-Request-ID", str(uuid4()))
    request.state.request_id = request_id
    response = await call_next(request)
    response.headers["X-Request-ID"] = request_id
    log_access(request, response.status_code, (time.perf_counter() - started) * 1000)
    return response


@app.exception_handler(HTTPException)
async def http_error(request: Request, exc: HTTPException):
    detail = exc.detail if isinstance(exc.detail, dict) else {"message": str(exc.detail)}
    log_request_error(request, exc.status_code, detail.get("code", "HTTP_ERROR"), detail.get("message", "Request failed."))
    return JSONResponse(status_code=exc.status_code, content={
        "error": {"code": detail.get("code", "HTTP_ERROR"), "message": detail.get("message", "Request failed."), **{key: value for key, value in detail.items() if key not in {"code", "message"}}},
        "request_id": getattr(request.state, "request_id", None),
    })


@app.exception_handler(RequestValidationError)
async def validation_error(request: Request, exc: RequestValidationError):
    fields = [{"field": ".".join(str(part) for part in item["loc"][1:]), "message": item["msg"]} for item in exc.errors()]
    log_request_error(request, 422, "VALIDATION_ERROR", f"Request validation failed for {len(fields)} field(s).")
    return JSONResponse(status_code=422, content={
        "error": {"code": "VALIDATION_ERROR", "message": "Please check the submitted values.", "fields": fields},
        "request_id": getattr(request.state, "request_id", None),
    })


@app.exception_handler(Exception)
async def unhandled_error(request: Request, exc: Exception):
    log_request_error(request, 500, type(exc).__name__, "Unhandled server error.", exc=exc)
    return JSONResponse(status_code=500, content={
        "error": {"code": "INTERNAL_SERVER_ERROR", "message": "An unexpected server error occurred."},
        "request_id": getattr(request.state, "request_id", None),
    })


PROJECT_FIELD_NAMES = list(ProjectCreate.model_fields)
SORT_COLUMNS = {
    "created_at": Project.created_at,
    "updated_at": Project.updated_at,
    "ceg": Project.ceg,
    "budget": Project.budget,
    "estimated_closing_date": Project.estimated_closing_date,
    "priority": Project.project_priority,
}
PRIORITY_ORDER = case(
    (Project.project_priority == "High", 1),
    (Project.project_priority == "Medium", 2),
    (Project.project_priority == "Normal", 3),
    else_=4,
)


def json_value(value):
    if isinstance(value, (date, Decimal)):
        return str(value)
    return value


def project_read(project: Project) -> ProjectRead:
    data = {column.name: getattr(project, column.name) for column in Project.__table__.columns}
    data["is_overdue"] = bool(
        project.lifecycle == Lifecycle.active.value
        and project.estimated_closing_date
        and project.estimated_closing_date < date.today()
    )
    data["project_cycle_business_days"] = (
        business_days_between(project.pr_approved_date, project.po_release_date)
        if project.pr_approved_date and project.po_release_date and project.po_release_date >= project.pr_approved_date
        else None
    )
    return ProjectRead.model_validate(data)


def add_audit(db: Session, project: Project, action: str, changes: dict, actor: Actor):
    db.add(ProjectAuditLog(
        project_id=project.id,
        action=action,
        changes_json=json.dumps(changes, ensure_ascii=False, default=json_value),
        actor_id=actor.id,
        actor_name=actor.name,
    ))


REFERENCE_PROJECT_FIELDS = {
    "supplier_type": "supplier_type",
    "procurement_strategy": "procurement_strategy",
    "procurement_status": "procurement_status",
}


def validate_reference_values(db: Session, values: dict, previous: Project | None = None):
    for field, category in REFERENCE_PROJECT_FIELDS.items():
        value = values.get(field)
        if not value or (previous is not None and getattr(previous, field) == value):
            continue
        exists = db.scalar(select(func.count()).select_from(ReferenceOption).where(
            ReferenceOption.category == category,
            ReferenceOption.code == value,
            ReferenceOption.active.is_(True),
        ))
        if not exists:
            raise HTTPException(422, detail={"code": "INVALID_OPTION", "message": f"{field} is not an active option."})


def query_projects(
    lifecycle: str | None = None,
    priority: str | None = None,
    ceg: str | None = None,
    keyword: str | None = None,
    procurement_status: str | None = None,
    bu: str | None = None,
    requestor: str | None = None,
    pr_approved_from: date | None = None,
    pr_approved_to: date | None = None,
    closing_from: date | None = None,
    closing_to: date | None = None,
    po_release_from: date | None = None,
    po_release_to: date | None = None,
    overdue: bool | None = None,
):
    filters = [Project.deleted_at.is_(None)]
    if lifecycle:
        filters.append(Project.lifecycle == lifecycle)
    if priority:
        filters.append(Project.project_priority == priority)
    if ceg:
        filters.append(func.lower(Project.ceg).contains(ceg.strip().lower()))
    if keyword:
        term = f"%{keyword.strip().lower()}%"
        filters.append(or_(
            func.lower(Project.ceg).like(term), func.lower(Project.requestor).like(term),
            func.lower(Project.bu).like(term), func.lower(Project.description).like(term),
            func.lower(Project.supplier_name).like(term), func.lower(Project.procurement_status_notes).like(term),
        ))
    if procurement_status:
        filters.append(Project.procurement_status == procurement_status)
    if bu:
        filters.append(func.lower(Project.bu).contains(bu.strip().lower()))
    if requestor:
        filters.append(func.lower(Project.requestor).contains(requestor.strip().lower()))
    if pr_approved_from:
        filters.append(Project.pr_approved_date >= pr_approved_from)
    if pr_approved_to:
        filters.append(Project.pr_approved_date <= pr_approved_to)
    if closing_from:
        filters.append(Project.estimated_closing_date >= closing_from)
    if closing_to:
        filters.append(Project.estimated_closing_date <= closing_to)
    if po_release_from:
        filters.append(Project.po_release_date >= po_release_from)
    if po_release_to:
        filters.append(Project.po_release_date <= po_release_to)
    if overdue is True:
        filters.extend([
            Project.lifecycle == Lifecycle.active.value,
            Project.estimated_closing_date < date.today(),
        ])
    elif overdue is False:
        filters.append(or_(
            Project.lifecycle != Lifecycle.active.value,
            Project.estimated_closing_date.is_(None),
            Project.estimated_closing_date >= date.today(),
        ))
    return filters


def commit_or_conflict(db: Session, ceg_message: str = "CEG already exists."):
    try:
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(status_code=409, detail={"code": "CEG_CONFLICT", "message": ceg_message}) from exc


def seed_reference_options(db: Session):
    defaults = {
        "supplier_type": [("Payment Only", "仅付款"), ("Simplified", "简化采购"), ("Sporadic", "零星采购"), ("Official", "正式供应商")],
        "procurement_strategy": [("Negotiation", "谈判"), ("Cost Comparison", "比价"), ("Bidding", "招标")],
        "procurement_status": [("Sourcing", "寻源"), ("Qualification", "资质审核"), ("Supplier Selection", "供应商选择"), ("Contract Review", "合同审核"), ("PO Release", "采购订单发布"), ("Others", "其他")],
    }
    for category, values in defaults.items():
        for index, (value, label_zh) in enumerate(values):
            exists = db.scalar(select(ReferenceOption.id).where(
                ReferenceOption.category == category,
                ReferenceOption.code == value,
            ))
            if not exists:
                db.add(ReferenceOption(category=category, code=value, label_en=value, label_zh=label_zh, sort_order=index))
    db.commit()


@app.get("/api/health")
def health():
    return {"status": "ok"}


@app.get("/api/session")
def session(actor: Actor = Depends(get_actor)):
    return {"actor": actor}


def parse_iam_token_response(response: httpx.Response) -> str:
    for header_name in ("Authorization", "X-Subject-Token", "X-Auth-Token"):
        token = response.headers.get(header_name, "").strip()
        if token:
            return token

    try:
        payload = response.json()
    except ValueError as exc:
        raise ValueError("IAM service returned no authentication token.") from exc

    candidates = []
    if isinstance(payload, dict):
        data = payload.get("data")
        attributes = data.get("attributes") if isinstance(data, dict) else None
        if isinstance(attributes, dict):
            candidates.extend((attributes.get("token"), attributes.get("access_token"), attributes.get("accessToken")))
        if isinstance(data, dict):
            candidates.extend((data.get("token"), data.get("access_token"), data.get("accessToken"), data.get("id")))
        candidates.extend((payload.get("token"), payload.get("access_token"), payload.get("accessToken")))

    token = next((item.strip() for item in candidates if isinstance(item, str) and item.strip()), "")
    if not token:
        raise ValueError("IAM service returned no authentication token.")
    return token


def parse_exchange_rate_response(
    payload: dict, from_currency: str, to_currency: str, rate_type: str,
) -> tuple[Decimal, datetime | None]:
    if str(payload.get("status")) != "200":
        raise ValueError("Exchange rate service returned an unsuccessful status.")
    data = payload.get("data")
    result = data.get("result") if isinstance(data, dict) else None
    if not isinstance(result, list) or not result:
        raise ValueError("Exchange rate service returned no quote.")

    expected_from = from_currency.upper()
    expected_to = to_currency.upper()
    expected_type = rate_type.upper()
    quote = next((item for item in result if isinstance(item, dict)
        and str(item.get("from_currency", "")).upper() == expected_from
        and str(item.get("to_currency", "")).upper() == expected_to
        and str(item.get("rate_type", "")).upper() == expected_type), None)
    if quote is None:
        raise ValueError("Exchange rate service returned no matching quote.")

    try:
        rate = Decimal(str(quote.get("rate_value")))
    except (InvalidOperation, TypeError, ValueError) as exc:
        raise ValueError("Exchange rate service returned an invalid rate.") from exc
    if not rate.is_finite() or rate <= 0:
        raise ValueError("Exchange rate service returned an invalid rate.")

    rate_date = quote.get("rate_date")
    quoted_at = None
    if rate_date:
        try:
            quoted_date = date.fromisoformat(str(rate_date).replace("/", "-"))
        except ValueError as exc:
            raise ValueError("Exchange rate service returned an invalid rate date.") from exc
        quoted_at = datetime(quoted_date.year, quoted_date.month, quoted_date.day, tzinfo=timezone.utc)
    return rate, quoted_at


@app.get("/api/exchange-rate")
def exchange_rate(request: Request, currency: Currency, actor: Actor = Depends(get_actor)):
    started = time.perf_counter()
    fetched_at = datetime.now(timezone.utc)
    if currency == "USD":
        log_operation("exchange_rate_lookup", actor.id, count=0, duration_ms=(time.perf_counter() - started) * 1000,
                      request_id=request.state.request_id, message="USD base exchange rate returned.")
        return {"currency": currency, "target_currency": "USD", "rate": "1.00000000", "fetched_at": fetched_at, "source": "USD base rate"}
    enterprise_id = settings.exchange_rate_iam_enterprise_id.strip()
    tenant_id = settings.exchange_rate_tenant_id.strip() or enterprise_id
    rate_type = settings.exchange_rate_rate_type.strip()
    iam_account = settings.exchange_rate_iam_account.strip()
    iam_secret = settings.exchange_rate_iam_secret.get_secret_value().strip()
    iam_project_id = settings.exchange_rate_iam_project_id.strip()
    if not all((tenant_id, rate_type, settings.exchange_rate_iam_token_url.strip(), iam_account, iam_secret,
                iam_project_id, enterprise_id)):
        log_operation("exchange_rate_lookup", actor.id, result="not_configured", count=0,
                      duration_ms=(time.perf_counter() - started) * 1000, request_id=request.state.request_id,
                      message=f"Exchange rate lookup failed; currency={currency}.")
        raise HTTPException(503, detail={
            "code": "EXCHANGE_RATE_NOT_CONFIGURED",
            "message": "Exchange rate service is not configured.",
        })
    request_payload = {
        "tenant_id": tenant_id,
        "curPage": "1",
        "pageSize": "20",
        "multi_rate_type_flag": "Y",
        "data": [{
            "from_currency": currency,
            "to_currency": "USD",
            "rate_type": rate_type,
            "start_date": fetched_at.strftime("%Y/%m/%d"),
        }],
    }
    token_payload = {
        "data": {
            "type": "token",
            "attributes": {
                "account": iam_account,
                "secret": iam_secret,
                "project": iam_project_id,
                "enterprise": enterprise_id,
            },
        },
    }
    token_response = None
    token_started = time.perf_counter()
    try:
        token_response = httpx.post(
            settings.exchange_rate_iam_token_url,
            headers={"Content-Type": "application/json"},
            json=token_payload,
            follow_redirects=True,
            timeout=settings.exchange_rate_timeout_seconds,
        )
        token_response.raise_for_status()
        authorization_token = parse_iam_token_response(token_response)
        log_integration_event(
            request_id=request.state.request_id,
            actor_id=actor.id,
            service="huawei_iam",
            operation="token_fetch",
            result="success",
            url=settings.exchange_rate_iam_token_url,
            status=token_response.status_code,
            duration_ms=(time.perf_counter() - token_started) * 1000,
            message="Dynamic IAM token obtained; sensitive response content omitted.",
        )
    except (httpx.HTTPError, ValueError) as exc:
        log_integration_event(
            request_id=request.state.request_id,
            actor_id=actor.id,
            service="huawei_iam",
            operation="token_fetch",
            result="failed",
            url=settings.exchange_rate_iam_token_url,
            status=token_response.status_code if token_response is not None else "-",
            duration_ms=(time.perf_counter() - token_started) * 1000,
            response=summarize_http_response(token_response),
            message="Failed to obtain dynamic IAM token.",
            exc=exc,
        )
        log_operation("exchange_rate_lookup", actor.id, result="failed:iam", count=0,
                      duration_ms=(time.perf_counter() - started) * 1000, request_id=request.state.request_id,
                      message=f"Exchange rate lookup failed during IAM authentication; currency={currency}.")
        raise HTTPException(502, detail={
            "code": "EXCHANGE_RATE_AUTH_UNAVAILABLE",
            "message": "Exchange rate authentication is temporarily unavailable.",
        }) from exc

    rate_response = None
    rate_started = time.perf_counter()
    try:
        rate_response = httpx.post(
            settings.exchange_rate_api_url,
            headers={"Authorization": authorization_token},
            json=request_payload,
            follow_redirects=True,
            timeout=settings.exchange_rate_timeout_seconds,
        )
        rate_response.raise_for_status()
        rate, quoted_at = parse_exchange_rate_response(rate_response.json(), currency, "USD", rate_type)
        log_integration_event(
            request_id=request.state.request_id,
            actor_id=actor.id,
            service="huawei_idata_finance",
            operation="exchange_rate_fetch",
            result="success",
            url=settings.exchange_rate_api_url,
            status=rate_response.status_code,
            duration_ms=(time.perf_counter() - rate_started) * 1000,
            message=f"Exchange rate received; currency={currency}; target=USD; rate_type={rate_type}.",
        )
    except (httpx.HTTPError, ValueError) as exc:
        log_integration_event(
            request_id=request.state.request_id,
            actor_id=actor.id,
            service="huawei_idata_finance",
            operation="exchange_rate_fetch",
            result="failed",
            url=settings.exchange_rate_api_url,
            status=rate_response.status_code if rate_response is not None else "-",
            duration_ms=(time.perf_counter() - rate_started) * 1000,
            response=summarize_http_response(rate_response),
            message=f"Failed to obtain exchange rate; currency={currency}; target=USD; rate_type={rate_type}.",
            exc=exc,
        )
        log_operation("exchange_rate_lookup", actor.id, result=f"failed:{type(exc).__name__}", count=0,
                      duration_ms=(time.perf_counter() - started) * 1000, request_id=request.state.request_id,
                      message=f"Exchange rate lookup failed; currency={currency}.")
        raise HTTPException(502, detail={
            "code": "EXCHANGE_RATE_UNAVAILABLE",
            "message": "Exchange rate service is temporarily unavailable.",
        }) from exc
    log_operation("exchange_rate_lookup", actor.id, count=0, duration_ms=(time.perf_counter() - started) * 1000,
                  request_id=request.state.request_id, message=f"Exchange rate lookup succeeded; currency={currency}.")
    return {
        "currency": currency,
        "target_currency": "USD",
        "rate": str(rate),
        "fetched_at": quoted_at or fetched_at,
        "source": "Huawei iData Finance",
    }


@app.get("/api/projects", response_model=PaginatedProjects)
def list_projects(
    page: int = Query(1, ge=1), page_size: int = Query(25, ge=1, le=100),
    lifecycle: str | None = None, priority: str | None = None, ceg: str | None = None,
    keyword: str | None = None, procurement_status: str | None = None,
    bu: str | None = None, requestor: str | None = None,
    pr_approved_from: date | None = None, pr_approved_to: date | None = None,
    closing_from: date | None = None, closing_to: date | None = None,
    po_release_from: date | None = None, po_release_to: date | None = None,
    overdue: bool | None = None, sort: str = "priority", direction: str = "desc",
    db: Session = Depends(get_db), _actor: Actor = Depends(get_actor),
):
    filters = query_projects(lifecycle, priority, ceg, keyword, procurement_status, bu, requestor, pr_approved_from, pr_approved_to, closing_from, closing_to, po_release_from, po_release_to, overdue)
    total = db.scalar(select(func.count()).select_from(Project).where(*filters)) or 0
    if sort == "priority":
        ordering = (PRIORITY_ORDER.asc(), Project.updated_at.desc(), Project.id.desc())
    else:
        sort_column = SORT_COLUMNS.get(sort, Project.created_at)
        order = sort_column.asc() if direction == "asc" else sort_column.desc()
        ordering = (order, Project.id.desc())
    items = db.scalars(select(Project).where(*filters).order_by(*ordering).offset((page - 1) * page_size).limit(page_size)).all()
    return PaginatedProjects(items=[project_read(item) for item in items], total=total, page=page, page_size=page_size, pages=max(1, math.ceil(total / page_size)))


@app.get("/api/projects/{project_id}", response_model=ProjectRead)
def get_project(project_id: int, db: Session = Depends(get_db), _actor: Actor = Depends(get_actor)):
    project = db.get(Project, project_id)
    if not project or project.deleted_at is not None:
        raise HTTPException(404, detail={"code": "NOT_FOUND", "message": "Project not found."})
    return project_read(project)


@app.post("/api/projects", response_model=ProjectRead, status_code=201)
def create_project(payload: ProjectCreate, request: Request, db: Session = Depends(get_db), actor: Actor = Depends(get_actor)):
    values = payload.model_dump()
    validate_reference_values(db, values)
    if values.get("po_release_date") is not None:
        values["lifecycle"] = Lifecycle.completed.value
        values["completed_at"] = utcnow()
    project = Project(**values, created_by=actor.id, updated_by=actor.id)
    try:
        db.add(project)
        db.flush()
        add_audit(db, project, "created", {key: {"before": None, "after": json_value(value)} for key, value in values.items() if value is not None}, actor)
        commit_or_conflict(db)
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(status_code=409, detail={"code": "CEG_CONFLICT", "message": "CEG already exists."}) from exc
    db.refresh(project)
    log_operation("project_create", actor.id, project_id=project.id, ceg=project.ceg, request_id=request.state.request_id)
    return project_read(project)


@app.put("/api/projects/{project_id}", response_model=ProjectRead)
def update_project(project_id: int, payload: ProjectUpdate, request: Request, db: Session = Depends(get_db), actor: Actor = Depends(get_actor)):
    project = db.get(Project, project_id)
    if not project or project.deleted_at is not None:
        raise HTTPException(404, detail={"code": "NOT_FOUND", "message": "Project not found."})
    if project.version != payload.version:
        raise HTTPException(409, detail={"code": "VERSION_CONFLICT", "message": "This project was changed by another user. Refresh and try again.", "current_version": project.version})
    changes = {}
    values = payload.model_dump(exclude={"version"})
    validate_reference_values(db, values, project)
    for key, value in values.items():
        before = getattr(project, key)
        if before != value:
            changes[key] = {"before": json_value(before), "after": json_value(value)}
    if values.get("po_release_date") is not None and project.lifecycle != Lifecycle.completed.value:
        values["lifecycle"] = Lifecycle.completed.value
        values["completed_at"] = utcnow()
        changes["lifecycle"] = {"before": project.lifecycle, "after": Lifecycle.completed.value}
    if changes:
        result = db.execute(update(Project).where(Project.id == project_id, Project.version == payload.version).values(
            **values, version=payload.version + 1, updated_by=actor.id, updated_at=utcnow(),
        ))
        if result.rowcount != 1:
            db.rollback()
            raise HTTPException(409, detail={"code": "VERSION_CONFLICT", "message": "This project was changed by another user. Refresh and try again."})
        add_audit(db, project, "updated", changes, actor)
        commit_or_conflict(db)
        db.expire_all()
        project = db.get(Project, project_id)
        log_operation("project_update", actor.id, project_id=project.id, ceg=project.ceg, request_id=request.state.request_id)
    return project_read(project)


def change_lifecycle(project_id: int, target: Lifecycle, action: str, payload: LifecycleRequest, db: Session, actor: Actor, request_id: str):
    project = db.get(Project, project_id)
    if not project or project.deleted_at is not None:
        raise HTTPException(404, detail={"code": "NOT_FOUND", "message": "Project not found."})
    if project.version != payload.version:
        raise HTTPException(409, detail={"code": "VERSION_CONFLICT", "message": "This project was changed by another user."})
    allowed = {"completed": {Lifecycle.active.value}, "reopened": {Lifecycle.completed.value}}
    if project.lifecycle not in allowed[action]:
        raise HTTPException(409, detail={"code": "INVALID_TRANSITION", "message": f"Cannot mark a {project.lifecycle} project as {target.value}."})
    before = project.lifecycle
    completed_at = utcnow() if target == Lifecycle.completed else (None if target == Lifecycle.active else project.completed_at)
    result = db.execute(update(Project).where(Project.id == project_id, Project.version == payload.version).values(
        lifecycle=target.value, completed_at=completed_at, archived_at=None,
        version=payload.version + 1, updated_by=actor.id, updated_at=utcnow(),
    ))
    if result.rowcount != 1:
        db.rollback()
        raise HTTPException(409, detail={"code": "VERSION_CONFLICT", "message": "This project was changed by another user."})
    add_audit(db, project, action, {"lifecycle": {"before": before, "after": target.value}}, actor)
    db.commit()
    db.expire_all()
    project = db.get(Project, project_id)
    log_operation(f"project_{action}", actor.id, project_id=project.id, ceg=project.ceg, request_id=request_id)
    return project_read(project)


@app.post("/api/projects/{project_id}/complete", response_model=ProjectRead)
def complete_project(project_id: int, payload: LifecycleRequest, request: Request, db: Session = Depends(get_db), actor: Actor = Depends(get_actor)):
    return change_lifecycle(project_id, Lifecycle.completed, "completed", payload, db, actor, request.state.request_id)


@app.post("/api/projects/{project_id}/reopen", response_model=ProjectRead)
def reopen_project(project_id: int, payload: LifecycleRequest, request: Request, db: Session = Depends(get_db), actor: Actor = Depends(get_actor)):
    return change_lifecycle(project_id, Lifecycle.active, "reopened", payload, db, actor, request.state.request_id)


@app.delete("/api/projects/{project_id}", status_code=204)
def delete_project(project_id: int, request: Request, version: int = Query(ge=1), db: Session = Depends(get_db), actor: Actor = Depends(get_actor)):
    project = db.get(Project, project_id)
    if not project or project.deleted_at is not None:
        raise HTTPException(404, detail={"code": "NOT_FOUND", "message": "Project not found."})
    if project.version != version:
        raise HTTPException(409, detail={"code": "VERSION_CONFLICT", "message": "This project changed before it could be deleted. Refresh and try again."})
    now = utcnow()
    project.deleted_at = now
    project.deleted_by = actor.id
    project.version += 1
    project.updated_by = actor.id
    project.updated_at = now
    add_audit(db, project, "trashed", {"deleted_at": {"before": None, "after": json_value(now)}}, actor)
    db.commit()
    log_operation("project_trash", actor.id, project_id=project.id, ceg=project.ceg, request_id=request.state.request_id)
    return Response(status_code=204)


@app.post("/api/projects/bulk-delete")
def bulk_delete_projects(payload: BulkProjectDeleteRequest, request: Request, db: Session = Depends(get_db), actor: Actor = Depends(get_actor)):
    requested = {item.id: item.version for item in payload.projects}
    projects = list(db.scalars(select(Project).where(Project.id.in_(requested))))
    found = {project.id: project for project in projects}
    missing = sorted(set(requested) - set(found))
    conflicts = sorted(project_id for project_id, project in found.items() if project.version != requested[project_id] or project.deleted_at is not None)
    if missing or conflicts:
        raise HTTPException(409, detail={"code": "BULK_DELETE_CONFLICT", "message": "Some selected projects changed or no longer exist. Refresh and try again.", "missing_ids": missing, "conflict_ids": conflicts})
    now = utcnow()
    for project in projects:
        project.deleted_at = now
        project.deleted_by = actor.id
        project.version += 1
        project.updated_by = actor.id
        project.updated_at = now
        add_audit(db, project, "trashed", {"deleted_at": {"before": None, "after": json_value(now)}}, actor)
    db.commit()
    log_operation("project_bulk_trash", actor.id, count=len(projects), request_id=request.state.request_id)
    return {"deleted": len(projects)}


@app.get("/api/recycle-bin", response_model=PaginatedProjects)
def list_recycle_bin(page: int = Query(1, ge=1), page_size: int = Query(25, ge=1, le=100), db: Session = Depends(get_db), _actor: Actor = Depends(get_actor)):
    filters = [Project.deleted_at.is_not(None)]
    total = db.scalar(select(func.count()).select_from(Project).where(*filters)) or 0
    items = db.scalars(select(Project).where(*filters).order_by(Project.deleted_at.desc(), Project.id.desc()).offset((page - 1) * page_size).limit(page_size)).all()
    return PaginatedProjects(items=[project_read(item) for item in items], total=total, page=page, page_size=page_size, pages=max(1, math.ceil(total / page_size)))


@app.post("/api/recycle-bin/restore")
def restore_projects(payload: BulkProjectDeleteRequest, request: Request, db: Session = Depends(get_db), actor: Actor = Depends(get_actor)):
    requested = {item.id: item.version for item in payload.projects}
    projects = list(db.scalars(select(Project).where(Project.id.in_(requested))))
    found = {project.id: project for project in projects}
    missing = sorted(set(requested) - set(found))
    conflicts = sorted(project_id for project_id, project in found.items() if project.version != requested[project_id] or project.deleted_at is None)
    if missing or conflicts:
        raise HTTPException(409, detail={"code": "RESTORE_CONFLICT", "message": "Some selected projects changed or are no longer in the recycle bin. Refresh and try again.", "missing_ids": missing, "conflict_ids": conflicts})
    now = utcnow()
    for project in projects:
        before = project.deleted_at
        project.deleted_at = None
        project.deleted_by = None
        project.version += 1
        project.updated_by = actor.id
        project.updated_at = now
        add_audit(db, project, "restored", {"deleted_at": {"before": json_value(before), "after": None}}, actor)
    db.commit()
    log_operation("project_restore", actor.id, count=len(projects), request_id=request.state.request_id)
    return {"restored": len(projects)}


@app.post("/api/recycle-bin/permanent-delete")
def permanently_delete_projects(payload: BulkProjectDeleteRequest, request: Request, db: Session = Depends(get_db), actor: Actor = Depends(get_actor)):
    requested = {item.id: item.version for item in payload.projects}
    projects = list(db.scalars(select(Project).where(Project.id.in_(requested))))
    found = {project.id: project for project in projects}
    missing = sorted(set(requested) - set(found))
    conflicts = sorted(project_id for project_id, project in found.items() if project.version != requested[project_id] or project.deleted_at is None)
    if missing or conflicts:
        raise HTTPException(409, detail={"code": "PERMANENT_DELETE_CONFLICT", "message": "Some selected projects changed or are no longer in the recycle bin. Refresh and try again.", "missing_ids": missing, "conflict_ids": conflicts})
    for project in projects:
        db.delete(project)
    db.commit()
    log_operation("project_permanent_delete", actor.id, count=len(projects), request_id=request.state.request_id)
    return {"deleted": len(projects)}


@app.get("/api/projects/{project_id}/audit-logs", response_model=list[AuditLogRead])
def audit_logs(project_id: int, db: Session = Depends(get_db), _actor: Actor = Depends(get_actor)):
    if not db.get(Project, project_id):
        raise HTTPException(404, detail={"code": "NOT_FOUND", "message": "Project not found."})
    logs = db.scalars(select(ProjectAuditLog).where(ProjectAuditLog.project_id == project_id).order_by(ProjectAuditLog.created_at.desc())).all()
    return [AuditLogRead(id=log.id, project_id=log.project_id, action=log.action, changes=json.loads(log.changes_json), actor_id=log.actor_id, actor_name=log.actor_name, created_at=log.created_at) for log in logs]


@app.get("/api/reference-options", response_model=list[ReferenceOptionRead])
def list_options(category: str | None = None, include_inactive: bool = False, db: Session = Depends(get_db), _actor: Actor = Depends(get_actor)):
    statement = select(ReferenceOption)
    if category:
        statement = statement.where(ReferenceOption.category == category)
    if not include_inactive:
        statement = statement.where(ReferenceOption.active.is_(True))
    return db.scalars(statement.order_by(ReferenceOption.category, ReferenceOption.sort_order, ReferenceOption.label_en)).all()


@app.post("/api/reference-options", response_model=ReferenceOptionRead, status_code=201)
def create_option(payload: ReferenceOptionCreate, db: Session = Depends(get_db), _actor: Actor = Depends(get_actor)):
    option = ReferenceOption(**payload.model_dump())
    db.add(option)
    try:
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(409, detail={"code": "OPTION_CONFLICT", "message": "Option code already exists."}) from exc
    db.refresh(option)
    return option


@app.put("/api/reference-options/{option_id}", response_model=ReferenceOptionRead)
def update_option(option_id: int, payload: ReferenceOptionUpdate, db: Session = Depends(get_db), _actor: Actor = Depends(get_actor)):
    option = db.get(ReferenceOption, option_id)
    if not option:
        raise HTTPException(404, detail={"code": "NOT_FOUND", "message": "Option not found."})
    for key, value in payload.model_dump().items():
        setattr(option, key, value)
    db.commit()
    db.refresh(option)
    return option


@app.delete("/api/reference-options/{option_id}", status_code=204)
def delete_option(option_id: int, db: Session = Depends(get_db), _actor: Actor = Depends(get_actor)):
    option = db.get(ReferenceOption, option_id)
    if not option:
        raise HTTPException(404, detail={"code": "NOT_FOUND", "message": "Option not found."})
    project_field = {
        "supplier_type": Project.supplier_type,
        "procurement_strategy": Project.procurement_strategy,
        "procurement_status": Project.procurement_status,
    }[option.category]
    usage_count = db.scalar(select(func.count()).select_from(Project).where(project_field == option.code)) or 0
    if usage_count:
        raise HTTPException(409, detail={
            "code": "OPTION_IN_USE",
            "message": f"This option is used by {usage_count} project(s). Deactivate it instead of deleting it.",
            "usage_count": usage_count,
        })
    db.delete(option)
    db.commit()
    return Response(status_code=204)


@app.get("/api/dashboard")
def dashboard(db: Session = Depends(get_db), _actor: Actor = Depends(get_actor)):
    active_filter = Project.deleted_at.is_(None)
    lifecycle_counts = dict(db.execute(select(Project.lifecycle, func.count()).where(active_filter).group_by(Project.lifecycle)).all())
    priority_counts = dict(db.execute(select(Project.project_priority, func.count()).where(active_filter, Project.project_priority.is_not(None)).group_by(Project.project_priority)).all())
    status_counts = dict(db.execute(select(Project.procurement_status, func.count()).where(active_filter, Project.procurement_status.is_not(None)).group_by(Project.procurement_status)).all())
    overdue = db.scalar(select(func.count()).select_from(Project).where(active_filter, Project.lifecycle == "active", Project.estimated_closing_date < date.today())) or 0
    total_budget = db.scalar(select(func.coalesce(func.sum(Project.usd_amount), 0)).where(
        active_filter, Project.lifecycle == Lifecycle.active.value,
    )) or 0
    ceg_name = func.coalesce(func.nullif(Project.ceg, ""), "Unassigned")
    ceg_amount = func.coalesce(func.sum(Project.usd_amount), 0)
    ceg_overview_rows = db.execute(
        select(ceg_name, func.count(), ceg_amount)
        .where(active_filter)
        .group_by(ceg_name)
        .order_by(ceg_amount.desc(), func.count().desc())
    ).all()
    ceg_overview = [{"ceg": ceg, "project_count": count, "usd_amount": str(amount)} for ceg, count, amount in ceg_overview_rows]
    return {"lifecycle": lifecycle_counts, "overdue": overdue, "total_budget": str(total_budget), "priority": priority_counts, "procurement_status": status_counts, "ceg_overview": ceg_overview}


def parse_month(value: str | None, end: bool = False) -> date | None:
    if value is None or value == "":
        return None
    if not re.fullmatch(r"\d{4}-(0[1-9]|1[0-2])", value):
        raise HTTPException(422, detail={"code": "INVALID_MONTH", "message": "Month must use YYYY-MM format."})
    year, month = map(int, value.split("-"))
    if end:
        next_month = date(year + (month == 12), 1 if month == 12 else month + 1, 1)
        return next_month - timedelta(days=1)
    return date(year, month, 1)


@app.get("/api/ceg-analysis")
def ceg_analysis(
    from_month: str | None = None, to_month: str | None = None,
    ceg: str | None = None, lifecycle: str | None = None,
    priority: str | None = None, bu: str | None = None,
    db: Session = Depends(get_db), _actor: Actor = Depends(get_actor),
):
    start = parse_month(from_month)
    end = parse_month(to_month, end=True)
    if start and end and start > end:
        raise HTTPException(422, detail={"code": "INVALID_MONTH_RANGE", "message": "From Month cannot be later than To Month."})
    if lifecycle and lifecycle not in {Lifecycle.active.value, Lifecycle.completed.value}:
        raise HTTPException(422, detail={"code": "INVALID_LIFECYCLE", "message": "Lifecycle must be active or completed."})
    if priority and priority not in {"High", "Medium", "Normal"}:
        raise HTTPException(422, detail={"code": "INVALID_PRIORITY", "message": "Priority must be High, Medium, or Normal."})

    filters = [Project.deleted_at.is_(None)]
    if start:
        filters.append(Project.pr_approved_date >= start)
    if end:
        filters.append(Project.pr_approved_date <= end)
    if ceg:
        filters.append(Project.ceg == ceg)
    if lifecycle:
        filters.append(Project.lifecycle == lifecycle)
    if priority:
        filters.append(Project.project_priority == priority)
    if bu:
        filters.append(Project.bu == bu)

    today = date.today()
    rows = db.execute(select(
        Project.ceg,
        func.count(),
        func.coalesce(func.sum(Project.usd_amount), 0),
        func.sum(case((Project.project_priority == "High", 1), else_=0)),
        func.sum(case((Project.project_priority == "Medium", 1), else_=0)),
        func.sum(case((Project.project_priority == "Normal", 1), else_=0)),
        func.sum(case((Project.lifecycle == Lifecycle.completed.value, 1), else_=0)),
        func.sum(case((
            (Project.lifecycle == Lifecycle.active.value)
            & Project.estimated_closing_date.is_not(None)
            & (Project.estimated_closing_date < today), 1), else_=0)),
    ).where(*filters).group_by(Project.ceg).order_by(func.coalesce(func.sum(Project.usd_amount), 0).desc(), func.count().desc())).all()

    items = [{
        "ceg": row[0] or "Unassigned", "project_count": row[1], "usd_amount": str(row[2]),
        "high_priority_count": row[3] or 0, "medium_priority_count": row[4] or 0,
        "normal_priority_count": row[5] or 0, "completed_count": row[6] or 0,
        "overdue_count": row[7] or 0,
    } for row in rows]
    option_filters = [Project.deleted_at.is_(None)]
    ceg_options = list(db.scalars(select(Project.ceg).where(*option_filters, Project.ceg.is_not(None), Project.ceg != "").distinct().order_by(Project.ceg)))
    bu_options = list(db.scalars(select(Project.bu).where(*option_filters, Project.bu.is_not(None), Project.bu != "").distinct().order_by(Project.bu)))
    return {
        "items": items,
        "totals": {
            "project_count": sum(item["project_count"] for item in items),
            "usd_amount": str(sum((Decimal(item["usd_amount"]) for item in items), Decimal("0"))),
            "high_priority_count": sum(item["high_priority_count"] for item in items),
        },
        "options": {"ceg": ceg_options, "bu": bu_options},
    }


@app.get("/api/budget-analysis")
def budget_analysis(from_month: str | None = None, to_month: str | None = None, db: Session = Depends(get_db), _actor: Actor = Depends(get_actor)):
    start = parse_month(from_month)
    end = parse_month(to_month, end=True)
    if start and end and start > end:
        raise HTTPException(422, detail={"code": "INVALID_MONTH_RANGE", "message": "From Month cannot be later than To Month."})
    filters = [Project.deleted_at.is_(None), Project.lifecycle == Lifecycle.active.value, Project.pr_approved_date.is_not(None)]
    if start:
        filters.append(Project.pr_approved_date >= start)
    if end:
        filters.append(Project.pr_approved_date <= end)
    month_key = func.strftime("%Y-%m", Project.pr_approved_date)
    rows = db.execute(select(month_key, func.coalesce(func.sum(Project.usd_amount), 0), func.count()).where(*filters).group_by(month_key).order_by(month_key)).all()
    values = {month: (amount, count) for month, amount, count in rows}
    month_names = list(values)
    if start and end:
        month_names = []
        cursor = start
        while cursor <= end:
            month_names.append(cursor.strftime("%Y-%m"))
            cursor = date(cursor.year + (cursor.month == 12), 1 if cursor.month == 12 else cursor.month + 1, 1)
    monthly = [{"month": month, "usd_amount": str(values.get(month, (Decimal("0"), 0))[0]), "project_count": values.get(month, (Decimal("0"), 0))[1]} for month in month_names]
    return {"from_month": from_month, "to_month": to_month, "monthly": monthly, "total_usd_amount": str(sum((Decimal(item["usd_amount"]) for item in monthly), Decimal("0"))), "project_count": sum(item["project_count"] for item in monthly)}


def business_days_between(start: date, end: date) -> int:
    """Count Monday-Friday dates in [start, end)."""
    current = start
    total = 0
    while current < end:
        if current.weekday() < 5:
            total += 1
        current += timedelta(days=1)
    return total


def report_change(current: Decimal | int, previous: Decimal | int) -> tuple[str, str]:
    current_value, previous_value = Decimal(current), Decimal(previous)
    if previous_value == 0:
        return ("New", "positive") if current_value > 0 else ("0.0%", "neutral")
    change = (current_value - previous_value) / previous_value * 100
    return f"{change:+.1f}%", "positive" if change > 0 else "negative" if change < 0 else "neutral"


def monthly_report_html(month: str, projects: list[Project], previous_projects: list[Project], ceg: str | None, language: str = "en") -> str:
    safe = lambda value: escape(str(value if value not in (None, "") else "—"))
    amount = lambda value: f"{Decimal(value or 0):,.2f}"
    labels = {
        "Project Tracking Monthly Report": "项目跟踪月度报告", "PR Approved": "PR 批准月份", "CEG: All": "CEG：全部", "Procurement": "采购",
        "No projects in this report period.": "本报告期间内没有项目。", "No data": "暂无数据", "No overdue projects in this report period.": "本报告期间内没有逾期项目。",
        "Print / Save PDF": "打印 / 保存 PDF", "Executive Summary": "执行摘要", "Projects": "项目数量", "Total USD Amount": "美元总金额", "USD Amount": "美元金额", "New": "新增",
        "High Priority": "高优先级", "Overdue": "已逾期", "Completed": "已完成", "vs previous month": "较上月", "CEG Analysis": "CEG 分析",
        "Projects by CEG": "各 CEG 项目数量", "USD Amount by CEG": "各 CEG 美元金额", "Priority Mix": "优先级分布", "Priority by CEG": "各 CEG 优先级", "High": "高", "Medium": "中", "Normal": "普通", "Total": "总计",
        "Overdue Attention": "逾期项目关注", "Project Details": "项目明细", "Description": "项目描述", "BU Requestor": "BU 申请人", "Supplier": "供应商", "Estimated Closing": "预计结束日期",
        "Overdue Days": "逾期天数", "Status": "状态", "Notes": "备注", "Priority": "优先级", "PR Approved Date": "PR 批准日期", "Lifecycle": "项目状态",
        "Active": "进行中", "Confidential internal report": "内部保密报告", "Data source: local reporting database": "数据来源：本地报告数据库",
        "Sourcing": "寻源", "Qualification": "资质审核", "Supplier Selection": "供应商选择", "Contract Review": "合同审核", "PO Release": "采购订单发布", "Others": "其他",
    }
    label = lambda value: labels.get(value, value) if language == "zh" else value
    total_amount = sum((Decimal(project.usd_amount or 0) for project in projects), Decimal("0"))
    previous_amount = sum((Decimal(project.usd_amount or 0) for project in previous_projects), Decimal("0"))
    count_change, count_tone = report_change(len(projects), len(previous_projects))
    amount_change, amount_tone = report_change(total_amount, previous_amount)
    today = date.today()
    overdue_projects = [project for project in projects if project.lifecycle == Lifecycle.active.value and project.estimated_closing_date and project.estimated_closing_date < today]
    completed_count = sum(project.lifecycle == Lifecycle.completed.value for project in projects)
    priority_counts = {name: sum(project.project_priority == name for project in projects) for name in ("High", "Medium", "Normal")}

    ceg_groups: dict[str, dict[str, Decimal | int]] = {}
    for project in projects:
        name = project.ceg or "Unassigned"
        group = ceg_groups.setdefault(name, {"count": 0, "amount": Decimal("0"), "high": 0, "medium": 0, "normal": 0})
        group["count"] += 1
        group["amount"] += Decimal(project.usd_amount or 0)
        key = (project.project_priority or "").lower()
        if key in {"high", "medium", "normal"}:
            group[key] += 1
    ceg_rows = sorted(ceg_groups.items(), key=lambda item: (-int(item[1]["high"]), -int(item[1]["medium"]), -int(item[1]["normal"]), item[0]))
    max_ceg_amount = max((Decimal(data["amount"]) for _, data in ceg_rows), default=Decimal("1")) or Decimal("1")
    max_ceg_count = max((int(data["count"]) for _, data in ceg_rows), default=1)

    priority_colors = {"high": "#b4233f", "medium": "#d6a15e", "normal": "#6f9e7c"}
    def priority_segments(data: dict[str, Decimal | int]) -> str:
        return "".join(
            f'<i title="{label(name.title())}: {int(data[name])}" style="display:grid;place-items:center;flex:{int(data[name])};min-width:{"24px" if int(data[name]) else "0"};color:#fff;background:{color};font-style:normal;font-size:9px;font-weight:bold">{int(data[name]) if int(data[name]) else ""}</i>'
            for name, color in priority_colors.items()
            if int(data[name])
        )
    ceg_priority_chart = "".join(
        f'''<div style="display:grid;grid-template-columns:120px 1fr 42px;align-items:center;gap:12px;margin:12px 0"><strong style="overflow:hidden;color:#344054;text-overflow:ellipsis;white-space:nowrap;font-size:10px">{safe(name)}</strong><div style="height:24px;overflow:hidden;border-radius:6px;background:#eef2f6"><div style="display:flex;width:{int(data["count"]) / max_ceg_count * 100:.2f}%;height:100%;overflow:hidden;border-radius:6px">{priority_segments(data)}</div></div><b style="color:#667085;text-align:right;font-size:10px">{data["count"]}</b></div>'''
        for name, data in ceg_rows
    ) or f'<p class="empty">{label("No projects in this report period.")}</p>'
    count_bars = "".join(
        f'''<div class="bar-row"><span>{safe(name)}</span><div class="track"><i style="width:{int(data["count"]) / max_ceg_count * 100:.2f}%"></i></div><b>{data["count"]}</b></div>'''
        for name, data in sorted(ceg_rows, key=lambda item: (-int(item[1]["count"]), item[0]))
    ) or f'<p class="empty">{label("No data")}</p>'
    amount_bars = "".join(
        f'''<div class="bar-row"><span>{safe(name)}</span><div class="track amount"><i style="width:{Decimal(data["amount"]) / max_ceg_amount * 100:.2f}%"></i></div><b>USD {amount(data["amount"])}</b></div>'''
        for name, data in sorted(ceg_rows, key=lambda item: (-Decimal(item[1]["amount"]), item[0]))
    ) or f'<p class="empty">{label("No data")}</p>'
    priority_text_colors = {"High": "#b4233f", "Medium": "#8a5a20", "Normal": "#3f7252"}
    overdue_rows = "".join(
        f'''<tr><td>{safe(project.ceg)}</td><td>{safe(project.description)}</td><td>{safe(project.bu)}</td><td>{safe(project.supplier_name)}</td><td>USD {amount(project.usd_amount)}</td><td>{safe(project.estimated_closing_date)}</td><td class="danger">{business_days_between(project.estimated_closing_date, today)}</td><td>{safe(label(project.procurement_status or ""))}</td><td>{safe(project.procurement_status_notes)}</td></tr>'''
        for project in sorted(overdue_projects, key=lambda item: item.estimated_closing_date or today)
    ) or f'<tr><td colspan="9" class="empty">{label("No overdue projects in this report period.")}</td></tr>'
    priority_rank = {"High": 0, "Medium": 1, "Normal": 2}
    detail_rows = "".join(
        f'''<tr><td><span class="priority {safe(project.project_priority).lower()}" style="color:{priority_text_colors.get(project.project_priority or '', '#475467')}">{label(project.project_priority or "")}</span></td><td>{safe(project.ceg)}</td><td>{safe(project.description)}</td><td>{safe(project.bu)}</td><td>{safe(project.supplier_name)}</td><td>USD {amount(project.usd_amount)}</td><td>{safe(label(project.procurement_status or ""))}</td><td>{safe(project.pr_approved_date)}</td><td>{safe(project.estimated_closing_date)}</td><td>{label((project.lifecycle or "").title())}</td></tr>'''
        for project in sorted(projects, key=lambda item: (priority_rank.get(item.project_priority or "", 3), item.ceg or "", item.id))
    ) or f'<tr><td colspan="10" class="empty">{label("No projects in this report period.")}</td></tr>'
    scope = " · ".join(part for part in [f"{label('PR Approved')}: {month}", f"CEG: {ceg}" if ceg else label("CEG: All")])

    return f'''<!doctype html><html lang="{'zh-CN' if language == 'zh' else 'en'}"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>{label("Project Tracking Monthly Report")} - {safe(month)}</title><style>
    @page{{size:A4 landscape;margin:12mm}}*{{box-sizing:border-box}}body{{margin:0;color:#172033;background:#edf2f8;font-family:Arial,"Microsoft YaHei",sans-serif;font-size:12px}}.report{{max-width:1180px;margin:24px auto;padding:0 38px 34px;overflow:hidden;border:1px solid #dce4ee;border-radius:18px;background:#fff;box-shadow:0 18px 48px #22324d1a}}header.report-head{{display:flex;align-items:flex-end;justify-content:space-between;gap:24px;margin:0 -38px;padding:30px 38px 28px;color:#fff;background:linear-gradient(120deg,#172b4d,#244d7c 62%,#2c6d8f)}}.brand{{margin-bottom:14px;color:#fff;font-size:16px;font-weight:bold;letter-spacing:.08em}}h1{{margin:0 0 9px;font-size:29px;letter-spacing:-.02em}}h2{{display:flex;align-items:center;gap:10px;margin:0 0 15px;color:#172b4d;font-size:17px}}h2:before{{width:4px;height:17px;border-radius:3px;background:#3478b8;content:""}}h3{{margin:0 0 15px;color:#243b5a;font-size:13px}}p{{margin:0;color:#667085}}.report-head p{{color:#dbe9f5}}.report-month{{min-width:104px;padding:10px 15px;border:1px solid #ffffff40;border-radius:20px;background:#ffffff18;text-align:center;font-size:14px;font-weight:bold;letter-spacing:.04em}}.toolbar{{max-width:1180px;margin:18px auto 0;text-align:right}}button{{padding:10px 17px;border:0;border-radius:8px;color:#fff;background:#244d7c;box-shadow:0 5px 14px #244d7c33;cursor:pointer}}section{{margin-top:30px;break-inside:avoid}}.metrics{{display:grid;grid-template-columns:repeat(5,1fr);gap:12px}}.metric{{position:relative;min-height:104px;padding:17px;border:1px solid #dfe7f0;border-radius:12px;background:#f8fafc}}.metric:before{{position:absolute;top:0;right:16px;left:16px;height:3px;border-radius:0 0 3px 3px;background:#527aa5;content:""}}.metric:nth-child(2):before{{background:#2f80a5}}.metric:nth-child(3):before{{background:#c24c5a}}.metric:nth-child(4):before{{background:#d06a61}}.metric:nth-child(5):before{{background:#2d8a67}}.metric span{{display:block;color:#667085;font-size:9px;font-weight:bold;text-transform:uppercase;letter-spacing:.05em}}.metric strong{{display:block;margin-top:20px;color:#172b4d;font-size:22px}}.metric small{{display:block;margin-top:7px;color:#667085}}.positive{{color:#087554!important}}.negative,.danger{{color:#b4233f!important;font-weight:bold}}.neutral{{color:#667085}}.charts{{display:grid;grid-template-columns:1fr 1fr;gap:18px}}.card{{padding:19px;border:1px solid #dfe7f0;border-radius:12px;background:linear-gradient(180deg,#fff,#f9fbfd)}}.bar-row{{display:grid;grid-template-columns:120px 1fr 110px;align-items:center;gap:10px;margin:11px 0}}.bar-row>span{{overflow:hidden;color:#344054;text-overflow:ellipsis;white-space:nowrap}}.bar-row>b{{color:#344054;font-size:10px;text-align:right}}.track{{height:11px;overflow:hidden;border-radius:6px;background:#e8eef5}}.track i{{display:block;height:100%;border-radius:6px;background:linear-gradient(90deg,#315b8a,#4e81b5)}}.track.amount i{{background:linear-gradient(90deg,#258199,#43a7af)}}.priority-bar{{display:flex;height:32px;overflow:hidden;border-radius:8px;background:#eef2f6}}.priority-bar i{{display:grid;place-items:center;color:#fff;font-style:normal;font-weight:bold}}.priority-bar .high{{background:#bd3f50}}.priority-bar .medium{{background:#d99834}}.priority-bar .normal{{color:#344054;background:#c9d2de}}.legend{{display:flex;gap:18px;margin-top:13px;color:#667085}}.table-scroll{{overflow:hidden;border:1px solid #dfe7f0;border-radius:12px}}table{{width:100%;border-collapse:collapse;background:#fff}}th,td{{padding:10px;border-bottom:1px solid #e8edf3;text-align:left;vertical-align:top}}tr:last-child td{{border-bottom:0}}tbody tr:nth-child(even){{background:#fafbfd}}th{{color:#52657a;background:#edf3f8;font-size:9px;text-transform:uppercase;letter-spacing:.045em}}td{{font-size:10px}}.priority{{padding:3px 7px;border-radius:10px;font-size:9px;font-weight:bold}}.priority.high{{color:#b4233f;background:#fff0f2}}.priority.medium{{color:#9a6700;background:#fff7df}}.priority.normal{{color:#475467;background:#f2f4f7}}.empty{{padding:28px;color:#98a2b3;text-align:center}}footer{{margin-top:30px;padding-top:13px;border-top:1px solid #e5eaf0;color:#98a2b3;font-size:9px;text-align:center}}@media(max-width:800px){{.report{{margin:0;padding:0 22px 22px;border-radius:0}}.metrics,.charts{{grid-template-columns:1fr 1fr}}header.report-head{{align-items:flex-start;flex-direction:column;margin:0 -22px;padding:25px 22px}}.table-scroll{{overflow:auto}}}}@media print{{body{{background:#fff}}.toolbar{{display:none}}.report{{max-width:none;margin:0;padding:0;border:0;box-shadow:none}}header.report-head{{margin:0 0 24px;padding:24px 28px}}section{{break-inside:avoid}}}}
    </style></head><body><div class="toolbar"><button onclick="window.print()">{label("Print / Save PDF")}</button></div><main class="report"><header class="report-head"><div><div class="brand">🍁 CARI {label("Procurement")}</div><h1>{label("Project Tracking Monthly Report")}</h1><p>{safe(scope)}</p></div><div class="report-month">{safe(month)}</div></header>
    <section><h2>{label("Executive Summary")}</h2><div class="metrics"><div class="metric"><span>{label("Projects")}</span><strong>{len(projects)}</strong><small class="{count_tone}">{label(count_change)} {label("vs previous month")}</small></div><div class="metric"><span>{label("Total USD Amount")}</span><strong>USD {amount(total_amount)}</strong><small class="{amount_tone}">{label(amount_change)} {label("vs previous month")}</small></div><div class="metric"><span>{label("High Priority")}</span><strong>{priority_counts["High"]}</strong></div><div class="metric"><span>{label("Overdue")}</span><strong class="{'danger' if overdue_projects else ''}">{len(overdue_projects)}</strong></div><div class="metric"><span>{label("Completed")}</span><strong>{completed_count}</strong></div></div></section>
    <section><h2>{label("CEG Analysis")}</h2><div class="charts"><div class="card"><h3>{label("Projects by CEG")}</h3>{count_bars}</div><div class="card"><h3>{label("USD Amount by CEG")}</h3>{amount_bars}</div></div><div class="card" style="margin-top:16px"><div style="display:flex;align-items:center;justify-content:space-between;gap:18px;margin-bottom:16px"><h3 style="margin:0">{label("Priority by CEG")}</h3><div class="legend" style="margin:0;font-weight:bold"><span style="color:#b4233f">{label("High")}: {priority_counts["High"]}</span><span style="color:#8a5a20">{label("Medium")}: {priority_counts["Medium"]}</span><span style="color:#3f7252">{label("Normal")}: {priority_counts["Normal"]}</span></div></div>{ceg_priority_chart}</div></section>
    <section><h2>{label("Overdue Attention")}</h2><div class="table-scroll"><table><thead><tr><th>CEG</th><th>{label("Description")}</th><th>BU</th><th>{label("Supplier")}</th><th>{label("USD Amount")}</th><th>{label("Estimated Closing")}</th><th>{label("Overdue Days")}</th><th>{label("Status")}</th><th>{label("Notes")}</th></tr></thead><tbody>{overdue_rows}</tbody></table></div></section>
    <section><h2>{label("Project Details")}</h2><div class="table-scroll"><table><thead><tr><th>{label("Priority")}</th><th>CEG</th><th>{label("Description")}</th><th>BU</th><th>{label("Supplier")}</th><th>{label("USD Amount")}</th><th>{label("Status")}</th><th>{label("PR Approved Date")}</th><th>{label("Estimated Closing")}</th><th>{label("Lifecycle")}</th></tr></thead><tbody>{detail_rows}</tbody></table></div></section><footer>CARI {label("Procurement")} · {label("Confidential internal report")} · {label("Data source: local reporting database")}</footer></main></body></html>'''


@app.get("/api/monthly-report.html")
def export_monthly_report(
    request: Request, month: str, ceg: str | None = None, language: str = "en",
    download: bool = False, db: Session = Depends(get_db), actor: Actor = Depends(get_actor),
):
    start = parse_month(month)
    end = parse_month(month, end=True)
    filters = [Project.deleted_at.is_(None), Project.pr_approved_date >= start, Project.pr_approved_date <= end]
    previous_end = start - timedelta(days=1)
    previous_start = date(previous_end.year, previous_end.month, 1)
    previous_filters = [Project.deleted_at.is_(None), Project.pr_approved_date >= previous_start, Project.pr_approved_date <= previous_end]
    if ceg:
        filters.append(Project.ceg == ceg)
        previous_filters.append(Project.ceg == ceg)
    projects = list(db.scalars(select(Project).where(*filters)))
    previous_projects = list(db.scalars(select(Project).where(*previous_filters)))
    if language not in {"en", "zh"}:
        raise HTTPException(422, detail={"code": "INVALID_LANGUAGE", "message": "Language must be en or zh."})
    content = monthly_report_html(month, projects, previous_projects, ceg, language)
    headers = {"Content-Disposition": f'{"attachment" if download else "inline"}; filename="Project_Tracking_Monthly_Report_{month}.html"'}
    log_operation("monthly_report_export" if download else "monthly_report_preview", actor.id, count=len(projects), message=f"month={month}; ceg={ceg or 'all'}", request_id=request.state.request_id)
    return Response(content=content, media_type="text/html", headers=headers)


EXPORT_COLUMNS = [
    ("Priority", "project_priority"), ("CEG", "ceg"), ("BU", "bu"), ("BU Requestor", "requestor"),
    ("Request Date", "request_date"), ("Budget (excl.tax)", "budget"), ("Currency", "currency"),
    ("Exchange Rate", "exchange_rate"), ("USD Amount", "usd_amount"), ("Description", "description"),
    ("Supplier Name", "supplier_name"), ("Supplier Type", "supplier_type"),
    ("Procurement Strategy", "procurement_strategy"), ("Procurement Status", "procurement_status"),
    ("Procurement Status Notes", "procurement_status_notes"),
    ("PR Approved Date", "pr_approved_date"), ("Estimated Project Closing Date", "estimated_closing_date"),
    ("EC Form", "ec_form"), ("Contract Required", "contract_required"), ("PO Release Date", "po_release_date"),
    ("Lifecycle", "lifecycle"), ("Overdue", "is_overdue"),
]
EXPORT_LABELS_ZH = {
    "Priority": "优先级", "CEG": "CEG", "BU": "业务部门", "BU Requestor": "BU 申请人",
    "Request Date": "申请日期", "Budget (excl.tax)": "预算（不含税）", "Currency": "币种",
    "Exchange Rate": "汇率", "USD Amount": "美元金额", "Description": "项目描述",
    "Supplier Name": "供应商名称", "Supplier Type": "供应商类型", "Procurement Strategy": "采购策略",
    "Procurement Status": "采购状态", "Procurement Status Notes": "采购状态备注", "PR Approved Date": "PR 批准日期",
    "Estimated Project Closing Date": "预计项目结束日期", "EC Form": "EC Form",
    "Contract Required": "是否需要签合同", "PO Release Date": "PO 发布日期",
    "Lifecycle": "生命周期", "Overdue": "是否逾期",
}


@app.get("/api/projects-export.xlsx")
def export_projects(
    request: Request,
    lifecycle: str | None = None, priority: str | None = None, ceg: str | None = None,
    keyword: str | None = None, procurement_status: str | None = None,
    bu: str | None = None, requestor: str | None = None,
    pr_approved_from: date | None = None, pr_approved_to: date | None = None,
    closing_from: date | None = None, closing_to: date | None = None,
    po_release_from: date | None = None, po_release_to: date | None = None, overdue: bool | None = None,
    language: str = "en",
    db: Session = Depends(get_db), actor: Actor = Depends(get_actor),
):
    started = time.perf_counter()
    filters = query_projects(lifecycle, priority, ceg, keyword, procurement_status, bu, requestor, pr_approved_from, pr_approved_to, closing_from, closing_to, po_release_from, po_release_to, overdue)
    projects = db.scalars(select(Project).where(*filters).order_by(PRIORITY_ORDER.asc(), Project.updated_at.desc(), Project.id.desc())).all()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Projects"
    sheet.append([EXPORT_LABELS_ZH.get(label, label) if language == "zh" else label for label, _ in EXPORT_COLUMNS])
    for project in projects:
        data = project_read(project).model_dump()
        data["is_overdue"] = "Overdue" if data["is_overdue"] else "Not Overdue"
        sheet.append([data.get(field) for _, field in EXPORT_COLUMNS])
    budget_column = next(index for index, (_, field) in enumerate(EXPORT_COLUMNS, start=1) if field == "budget")
    usd_column = next(index for index, (_, field) in enumerate(EXPORT_COLUMNS, start=1) if field == "usd_amount")
    for row in range(2, sheet.max_row + 1):
        sheet.cell(row=row, column=budget_column).number_format = "#,##0.00"
        sheet.cell(row=row, column=usd_column).number_format = "#,##0.00"
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    analysis = workbook.create_sheet("Monthly Analysis")
    analysis.append(["Month", "Project Count", "Monthly Amount", "MoM Change"])
    monthly_values: dict[str, dict[str, Decimal | int]] = {}
    for project in projects:
        if not project.pr_approved_date:
            continue
        month = project.pr_approved_date.strftime("%Y-%m")
        entry = monthly_values.setdefault(month, {"count": 0, "amount": Decimal("0")})
        entry["count"] = int(entry["count"]) + 1
        entry["amount"] = Decimal(entry["amount"]) + (project.usd_amount or Decimal("0"))
    month_names = sorted(monthly_values)
    if pr_approved_from and pr_approved_to:
        month_names = []
        cursor = pr_approved_from.replace(day=1)
        final_month = pr_approved_to.replace(day=1)
        while cursor <= final_month:
            month_names.append(cursor.strftime("%Y-%m"))
            cursor = date(cursor.year + (cursor.month == 12), 1 if cursor.month == 12 else cursor.month + 1, 1)
    for month in month_names:
        entry = monthly_values.get(month, {"count": 0, "amount": Decimal("0")})
        analysis.append([month, entry["count"], entry["amount"]])
    for cell in analysis[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2C2C2E")
    monthly_last_row = analysis.max_row
    for row in range(2, monthly_last_row + 1):
        analysis.cell(row=row, column=3).number_format = '"USD "#,##0.00'
        if row == 2:
            analysis.cell(row=row, column=4, value="")
        else:
            analysis.cell(row=row, column=4, value=f'=IF(C{row-1}=0,IF(C{row}=0,0,1),C{row}/C{row-1}-1)')
            analysis.cell(row=row, column=4).number_format = "0.0%"
    analysis.column_dimensions["A"].width = 14
    analysis.column_dimensions["B"].width = 16
    analysis.column_dimensions["C"].width = 20
    analysis.column_dimensions["D"].width = 16
    analysis.freeze_panes = "A2"
    if monthly_last_row >= 2:
        categories = Reference(analysis, min_col=1, min_row=2, max_row=monthly_last_row)
        amounts = Reference(analysis, min_col=3, min_row=1, max_row=monthly_last_row)
        bars = BarChart()
        bars.type = "col"
        bars.add_data(amounts, titles_from_data=True)
        bars.set_categories(categories)
        month_category_ref = f"'Monthly Analysis'!$A$2:$A${monthly_last_row}"
        bars.series[0].cat = AxDataSource(strRef=StrRef(
            f=month_category_ref,
            strCache=StrData(ptCount=len(month_names), pt=[StrVal(idx=index, v=month) for index, month in enumerate(month_names)]),
        ))
        bars.title = "Monthly Amount"
        bars.y_axis.title = "Amount (USD)"
        bars.y_axis.numFmt = '"USD "#,##0.00'
        bars.y_axis.scaling.min = 0
        bars.x_axis.title = "Year-Month (YYYY-MM)"
        bars.x_axis.tickLblSkip = 1
        bars.x_axis.tickMarkSkip = 1
        bars.x_axis.tickLblPos = "low"
        bars.x_axis.noMultiLvlLbl = True
        bars.y_axis.tickLblPos = "nextTo"
        bars.x_axis.spPr = GraphicalProperties(ln=LineProperties(solidFill="595959", w=12700))
        bars.y_axis.spPr = GraphicalProperties(ln=LineProperties(solidFill="595959", w=12700))
        bars.y_axis.majorGridlines = ChartLines(spPr=GraphicalProperties(ln=LineProperties(solidFill="D9E2F3", w=6350)))
        bars.legend = None
        bars.height = 11.5
        bars.width = 28
        bars.gapWidth = 65
        bars.series[0].graphicalProperties.solidFill = "5B9BD5"
        bars.series[0].graphicalProperties.line.solidFill = "4472C4"
        bars.dLbls = DataLabelList()
        bars.dLbls.showVal = True
        bars.dLbls.numFmt = '"USD "#,##0.00'
        bars.dLbls.dLblPos = "inEnd"
        analysis.add_chart(bars, "F2")
        line = LineChart()
        monthly_change = Reference(analysis, min_col=4, min_row=1, max_row=monthly_last_row)
        line.add_data(monthly_change, titles_from_data=True)
        line.set_categories(categories)
        line.series[0].cat = AxDataSource(strRef=StrRef(
            f=month_category_ref,
            strCache=StrData(ptCount=len(month_names), pt=[StrVal(idx=index, v=month) for index, month in enumerate(month_names)]),
        ))
        line.y_axis.title = "Change (%)"
        line.y_axis.numFmt = "0%"
        line.y_axis.tickLblPos = "nextTo"
        line.x_axis.tickLblPos = "low"
        line.x_axis.spPr = GraphicalProperties(ln=LineProperties(solidFill="595959", w=12700))
        line.y_axis.spPr = GraphicalProperties(ln=LineProperties(solidFill="595959", w=12700))
        line.y_axis.majorGridlines = ChartLines(spPr=GraphicalProperties(ln=LineProperties(solidFill="FCE4D6", w=6350)))
        line.title = "Monthly Change"
        line.x_axis.title = "Year-Month (YYYY-MM)"
        line.legend = None
        line.x_axis.tickLblSkip = 1
        line.x_axis.tickMarkSkip = 1
        line.x_axis.noMultiLvlLbl = True
        line.height = 11.5
        line.width = 28
        line.series[0].graphicalProperties.line.solidFill = "ED7D31"
        line.series[0].graphicalProperties.line.width = 19050
        line.series[0].marker.symbol = "circle"
        line.series[0].marker.size = 7
        line.series[0].marker.graphicalProperties.solidFill = "ED7D31"
        line.series[0].marker.graphicalProperties.line.solidFill = "C65911"
        line.dLbls = DataLabelList()
        line.dLbls.showVal = True
        line.dLbls.numFmt = "0.0%"
        line.dLbls.dLblPos = "t"
        analysis.add_chart(line, "F28")
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    log_operation("project_export", actor.id, count=len(projects), duration_ms=(time.perf_counter() - started) * 1000,
                  request_id=request.state.request_id, message="Project Excel export generated.")
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=procurement-projects.xlsx"})
