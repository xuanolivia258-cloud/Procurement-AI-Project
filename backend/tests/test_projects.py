from pathlib import Path

from app.config import settings


def test_backend_errors_are_logged_with_request_context(client):
    response = client.get("/api/projects/not-a-number", headers={"X-Request-ID": "test-request-id"})
    assert response.status_code == 422
    log_text = Path(settings.error_log_file).read_text(encoding="utf-8")
    assert "request_id=test-request-id" in log_text
    assert "method=GET" in log_text
    assert "path=/api/projects/not-a-number" in log_text
    assert "status=422" in log_text
    assert "error_type=VALIDATION_ERROR" in log_text


def test_access_and_project_operations_are_logged(client):
    response = client.post("/api/projects", json={"ceg": "LOGGED-PROJECT"}, headers={"X-Request-ID": "operation-log-test"})
    assert response.status_code == 201

    access_log = Path(settings.access_log_file).read_text(encoding="utf-8")
    assert "request_id=operation-log-test" in access_log
    assert "actor_id=local-test-user" in access_log
    assert "method=POST" in access_log
    assert "path=/api/projects" in access_log
    assert "status=201" in access_log
    assert "duration_ms=" in access_log

    operation_log = Path(settings.operation_log_file).read_text(encoding="utf-8")
    assert "request_id=operation-log-test" in operation_log
    assert "action=project_create" in operation_log
    assert f"project_id={response.json()['id']}" in operation_log
    assert "ceg=LOGGED-PROJECT" in operation_log


def test_empty_project_can_be_created(client):
    response = client.post("/api/projects", json={})
    assert response.status_code == 201
    project = response.json()
    assert project["lifecycle"] == "active"
    assert project["version"] == 1


def test_pr_approved_without_po_release_can_be_saved(client):
    response = client.post("/api/projects", json={"pr_approved_date": "2026-08-11"})
    assert response.status_code == 201
    assert response.json()["project_cycle_business_days"] is None
    assert client.get("/api/projects").status_code == 200


def test_po_release_date_automatically_completes_new_project(client):
    response = client.post("/api/projects", json={"ceg": "AUTO-COMPLETE-NEW", "po_release_date": "2026-08-14"})
    assert response.status_code == 201
    project = response.json()
    assert project["lifecycle"] == "completed"
    assert project["completed_at"] is not None
    assert client.get("/api/projects?lifecycle=completed&keyword=AUTO-COMPLETE-NEW").json()["total"] == 1
    assert client.get("/api/projects?lifecycle=active&keyword=AUTO-COMPLETE-NEW").json()["total"] == 0


def test_adding_po_release_date_automatically_completes_existing_project(client):
    project = client.post("/api/projects", json={"ceg": "AUTO-COMPLETE-EDIT"}).json()
    response = client.put(f"/api/projects/{project['id']}", json={
        "ceg": project["ceg"],
        "po_release_date": "2026-08-14",
        "version": project["version"],
    })
    assert response.status_code == 200
    updated = response.json()
    assert updated["lifecycle"] == "completed"
    assert updated["completed_at"] is not None
    logs = client.get(f"/api/projects/{project['id']}/audit-logs").json()
    assert logs[0]["changes"]["lifecycle"] == {"before": "active", "after": "completed"}


def test_free_text_starts_with_uppercase_english_letter(client):
    response = client.post("/api/projects", json={
        "ceg": "project-1", "bu": "finance", "requestor": "alice",
        "description": "new sourcing project", "supplier_name": "example supplier",
        "procurement_status_notes": "waiting for supplier feedback",
    })
    assert response.status_code == 201
    project = response.json()
    assert project["ceg"] == "Project-1"
    assert project["bu"] == "Finance"
    assert project["requestor"] == "Alice"
    assert project["description"] == "New sourcing project"
    assert project["supplier_name"] == "Example supplier"
    assert project["procurement_status_notes"] == "Waiting for supplier feedback"


def test_contract_required_accepts_only_supported_options(client):
    response = client.post("/api/projects", json={"contract_required": "N/A"})
    assert response.status_code == 201
    assert response.json()["contract_required"] == "N/A"

    invalid = client.post("/api/projects", json={"contract_required": "Maybe"})
    assert invalid.status_code == 422


def test_ec_form_accepts_na(client):
    response = client.post("/api/projects", json={"ec_form": "N/A"})
    assert response.status_code == 201
    assert response.json()["ec_form"] == "N/A"


def test_project_accepts_supported_currency(client):
    response = client.post("/api/projects", json={"budget": "123.45", "currency": "CAD", "exchange_rate": "0.75"})
    assert response.status_code == 201
    assert response.json()["currency"] == "CAD"
    assert response.json()["usd_amount"] == "92.59"
    assert client.post("/api/projects", json={"currency": "EUR"}).status_code == 422


def test_ceg_can_be_reused_when_copying_a_project(client):
    assert client.post("/api/projects", json={"ceg": "CEG-100"}).status_code == 201
    response = client.post("/api/projects", json={"ceg": "ceg-100"})
    assert response.status_code == 201
    assert response.json()["ceg"] == "Ceg-100"


def test_projects_default_to_business_priority_order(client):
    client.post("/api/projects", json={"ceg": "NORMAL-SORT", "project_priority": "Normal"})
    client.post("/api/projects", json={"ceg": "HIGH-SORT", "project_priority": "High"})
    client.post("/api/projects", json={"ceg": "EMPTY-SORT"})
    client.post("/api/projects", json={"ceg": "MEDIUM-SORT", "project_priority": "Medium"})

    items = client.get("/api/projects?lifecycle=active").json()["items"]
    assert [item["project_priority"] for item in items] == ["High", "Medium", "Normal", None]


def test_optimistic_lock_prevents_lost_update(client):
    project = client.post("/api/projects", json={"ceg": "LOCK-1"}).json()
    payload = {"ceg": "LOCK-1", "requestor": "First", "version": project["version"]}
    updated = client.put(f"/api/projects/{project['id']}", json=payload)
    assert updated.status_code == 200
    assert updated.json()["version"] == 2
    conflict = client.put(f"/api/projects/{project['id']}", json={**payload, "requestor": "Second"})
    assert conflict.status_code == 409
    assert conflict.json()["error"]["code"] == "VERSION_CONFLICT"


def test_lifecycle_and_audit_history(client):
    project = client.post("/api/projects", json={"ceg": "LIFE-1"}).json()
    completed = client.post(f"/api/projects/{project['id']}/complete", json={"version": project["version"]})
    assert completed.status_code == 200
    assert completed.json()["lifecycle"] == "completed"
    logs = client.get(f"/api/projects/{project['id']}/audit-logs").json()
    assert [entry["action"] for entry in logs] == ["completed", "created"]
    assert client.post(f"/api/projects/{project['id']}/archive", json={"version": 2}).status_code == 404


def test_project_can_be_deleted_with_current_version(client):
    project = client.post("/api/projects", json={"ceg": "DELETE-1"}).json()
    stale = client.delete(f"/api/projects/{project['id']}?version=2")
    assert stale.status_code == 409

    deleted = client.delete(f"/api/projects/{project['id']}?version={project['version']}")
    assert deleted.status_code == 204
    assert client.get(f"/api/projects/{project['id']}").status_code == 404
    recycled = client.get("/api/recycle-bin").json()
    assert recycled["total"] == 1
    assert recycled["items"][0]["ceg"] == "DELETE-1"

    restored = client.post("/api/recycle-bin/restore", json={"projects": [{"id": recycled["items"][0]["id"], "version": recycled["items"][0]["version"]}]})
    assert restored.json() == {"restored": 1}
    assert client.get("/api/projects?keyword=DELETE-1").json()["total"] == 1


def test_projects_can_be_bulk_deleted(client):
    first = client.post("/api/projects", json={"ceg": "BULK-1"}).json()
    second = client.post("/api/projects", json={"ceg": "BULK-2"}).json()
    response = client.post("/api/projects/bulk-delete", json={"projects": [{"id": first["id"], "version": first["version"]}, {"id": second["id"], "version": second["version"]}]})
    assert response.status_code == 200
    assert response.json() == {"deleted": 2}
    assert client.get("/api/projects?keyword=BULK-").json()["total"] == 0
    recycled = client.get("/api/recycle-bin").json()
    permanent = client.post("/api/recycle-bin/permanent-delete", json={"projects": [{"id": item["id"], "version": item["version"]} for item in recycled["items"]]})
    assert permanent.json() == {"deleted": 2}
    assert client.get("/api/recycle-bin").json()["total"] == 0


def test_filters_dashboard_and_excel_export(client):
    client.post("/api/projects", json={"ceg": "HIGH-1", "project_priority": "High", "budget": "125.50", "pr_approved_date": "2026-08-07", "po_release_date": "2026-08-11"})
    client.post("/api/projects", json={"ceg": "NORMAL-1", "project_priority": "Normal"})
    result = client.get("/api/projects?priority=High").json()
    assert result["total"] == 1
    assert result["items"][0]["ceg"] == "HIGH-1"
    dashboard = client.get("/api/dashboard").json()
    assert dashboard["priority"]["High"] == 1
    project = client.get("/api/projects?priority=High").json()["items"][0]
    assert project["project_cycle_business_days"] == 2
    assert client.get("/api/projects?pr_approved_from=2026-08-08").json()["total"] == 0
    assert client.get("/api/projects?pr_approved_from=2026-08-07&pr_approved_to=2026-08-07").json()["total"] == 1
    assert "average_project_cycle_business_days" not in dashboard
    export = client.get("/api/projects-export.xlsx?priority=High")
    assert export.status_code == 200
    assert export.headers["content-type"].startswith("application/vnd.openxmlformats")
    from io import BytesIO
    from openpyxl import load_workbook
    sheet = load_workbook(BytesIO(export.content)).active
    overdue_column = [cell.value for cell in sheet[1]].index("Overdue") + 1
    assert sheet.cell(row=2, column=overdue_column).value == "Not Overdue"
    workbook = load_workbook(BytesIO(export.content))
    assert workbook.sheetnames == ["Projects", "Monthly Analysis"]
    assert workbook["Monthly Analysis"]["A1"].value == "Month"
    assert [workbook["Monthly Analysis"].cell(1, column).value for column in range(1, 5)] == ["Month", "Project Count", "Monthly Amount", "MoM Change"]
    assert len(workbook["Monthly Analysis"]._charts) == 2
    assert workbook["Monthly Analysis"]["A2"].value == "2026-08"
    amount_chart, change_chart = workbook["Monthly Analysis"]._charts
    assert amount_chart.series[0].cat.strRef.f == "'Monthly Analysis'!$A$2:$A$2"
    assert [point.v for point in amount_chart.series[0].cat.strRef.strCache.pt] == ["2026-08"]
    assert amount_chart.x_axis.tickLblSkip == 1
    assert amount_chart.x_axis.noMultiLvlLbl is True
    assert amount_chart.anchor.ext.width == 28 * 360000
    assert amount_chart.anchor.ext.height == 11.5 * 360000
    assert amount_chart.y_axis.numFmt.formatCode == '"USD "#,##0.00'
    assert amount_chart.dLbls.showVal is True
    assert amount_chart.dLbls.numFmt == '"USD "#,##0.00'
    assert amount_chart.dLbls.dLblPos == "inEnd"
    assert amount_chart.x_axis.tickLblPos == "low"
    assert amount_chart.series[0].graphicalProperties.solidFill.srgbClr == "5B9BD5"
    assert change_chart.y_axis.numFmt.formatCode == "0%"
    assert change_chart.series[0].cat.strRef.f == "'Monthly Analysis'!$A$2:$A$2"
    assert [point.v for point in change_chart.series[0].cat.strRef.strCache.pt] == ["2026-08"]
    assert change_chart.anchor.ext.width == 28 * 360000
    assert change_chart.anchor.ext.height == 11.5 * 360000
    assert change_chart.dLbls.showVal is True
    assert change_chart.dLbls.numFmt == "0.0%"
    assert change_chart.series[0].graphicalProperties.line.solidFill.srgbClr == "ED7D31"
    assert change_chart.series[0].marker.symbol == "circle"
    assert change_chart.x_axis.tickLblPos == "low"


def test_dashboard_total_budget_sums_project_usd_amounts(client):
    client.post("/api/projects", json={"ceg": "CAD-1", "budget": "100.00", "currency": "CAD", "usd_amount": "72.50"})
    client.post("/api/projects", json={"ceg": "CNY-1", "budget": "1000.00", "currency": "CNY", "usd_amount": "138.25"})
    client.post("/api/projects", json={"ceg": "NO-USD", "budget": "999.00", "currency": "CAD"})
    client.post("/api/projects", json={"ceg": "COMPLETED-USD", "budget": "500.00", "currency": "USD", "exchange_rate": "1", "po_release_date": "2026-08-14"})

    dashboard = client.get("/api/dashboard").json()

    assert dashboard["total_budget"] == "210.75"


def test_budget_analysis_groups_usd_amount_by_pr_approved_month(client):
    client.post("/api/projects", json={"ceg": "JAN-1", "pr_approved_date": "2026-01-15", "budget": "100.25", "currency": "USD", "exchange_rate": "1"})
    client.post("/api/projects", json={"ceg": "MAR-1", "pr_approved_date": "2026-03-02", "budget": "50.00", "currency": "USD", "exchange_rate": "1"})

    result = client.get("/api/budget-analysis?from_month=2026-01&to_month=2026-03").json()

    assert result["monthly"] == [
        {"month": "2026-01", "usd_amount": "100.25", "project_count": 1},
        {"month": "2026-02", "usd_amount": "0", "project_count": 0},
        {"month": "2026-03", "usd_amount": "50.00", "project_count": 1},
    ]
    assert result["total_usd_amount"] == "150.25"
